import pandas as pd
import numpy as np
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

# Create a Tkinter root window
root = Tk()
# Hide the root window
root.withdraw()

try:
    # Open the file picker dialog
    clockIn_File = askopenfilename()
    payRoll_File = askopenfilename()

    # Check if a file path was selected
    if not clockIn_File or not payRoll_File:
        print("No file selected.")
        raise SystemExit

    # Read the Excel files
    df1 = pd.read_excel(clockIn_File)
    df2 = pd.read_excel(payRoll_File)

    # Keep only 'Employee name', 'Ticket Date' and 'Agency' columns in df2
    df2 = df2[
        ['Employee Name', 'Employee ID', 'Ticket Date', 'Agency', 'Clock-In ID', 'Supervisors Name', 'PM Assigned',
         'JobNo|Customer|Description', 'WTL Approved', 'WTL Start Date', 'WTL End Date', 'ApprovedOvertime',
         'ApprovedOvertime Start Date', 'ApprovedOvertime End Date']]

    # Convert 'Ticket Date' to datetime in both dataframes
    df1['Ticket Date'] = pd.to_datetime(df1['Ticket Date'])
    df2['Ticket Date'] = pd.to_datetime(df2['Ticket Date'])

    # Calculate 'Lunch Adjusted' as the difference between 'Clock Out' and 'Clock In', converted to hours
    df1['Lunch Adjusted'] = (
                                    df1['Clock Out'] - df1['Clock In']).dt.total_seconds() / 3600

    # Taking off the half hour for lunch if Hours Worked is greater than or equal to 5
    df1.loc[df1['Hours Worked'] >= 5, 'Lunch Adjusted'] -= 0.5

    # Add 'Day of the Week' column
    df1['Day of the Week'] = df1['Ticket Date'].dt.day_name()

    # Merge dataframes based on 'Employee name' and 'Ticket Date'
    merged_df = pd.merge(df1, df2, on=[
        'Employee Name', 'Ticket Date', 'JobNo|Customer|Description'], how='left')

    # Remove duplicates from merged_df based on 'Employee name', 'Ticket Date', and 'JobNo|Customer|Description'
    merged_df = merged_df.drop_duplicates(
        ['Employee Name', 'Ticket Date', 'JobNo|Customer|Description'])

    # # If 'Agency' is blank, fill with 'CSI'
    # merged_df['Agency'] = merged_df['Agency'].fillna('CSI')

    # Calculate Regular Time
    merged_df['Regular Time'] = merged_df['Lunch Adjusted'].where(
        merged_df['Lunch Adjusted'] <= 8, other=8)

    # Add 0.5 to 'Lunch Adjusted' column if there is a WTL Start Date and WTL End Date
    merged_df.loc[
        ~merged_df['WTL Start Date'].isnull() & ~merged_df['WTL End Date'].isnull(), 'Lunch Adjusted'] += 0.5

    # Calculate Overtime
    merged_df['Overtime'] = 0  # merged_df['Lunch Adjusted'] - 8

    # Calculate the overtime for each ticket, considering that the same employee
    # can have more than one ticket in a day.
    grouped_by_name = merged_df.groupby(
        ['Employee Name'])

    for name, group_name in grouped_by_name:
        grouped_by_name_by_date = group_name.groupby(
            [merged_df['Ticket Date'].dt.date])
        for date, indices in grouped_by_name_by_date.groups.items():
            worked_hours_needed = 8
            for index in indices:
                if worked_hours_needed == 0:
                    merged_df.loc[index,
                    'Overtime'] = merged_df.loc[index, 'Lunch Adjusted']
                elif worked_hours_needed >= merged_df.loc[index, 'Lunch Adjusted']:
                    worked_hours_needed -= merged_df.loc[index,
                    'Lunch Adjusted']
                else:
                    merged_df.loc[index, 'Overtime'] = merged_df.loc[index,
                    'Lunch Adjusted'] - worked_hours_needed
                    worked_hours_needed = 0
    merged_weekly_df = merged_df.copy(deep=True)
    # Calculate Overtime
    merged_weekly_df['Overtime'] = 0
    grouped_by_name = merged_weekly_df.groupby(
        ['Employee Name'])
    for name, group_name in grouped_by_name:
        grouped_by_name_by_date = group_name.groupby(
            [merged_weekly_df['Ticket Date'].dt.date])
        worked_hours_needed = 40
        for date, indices in grouped_by_name_by_date.groups.items():
            for index in indices:
                if worked_hours_needed == 0:
                    merged_weekly_df.loc[index, 'Overtime'] = merged_weekly_df.loc[index, 'Lunch Adjusted']
                    merged_weekly_df.loc[index, 'Regular Time'] = 0
                elif worked_hours_needed >= merged_weekly_df.loc[index, 'Lunch Adjusted']:
                    worked_hours_needed -= merged_weekly_df.loc[index, 'Lunch Adjusted']
                    merged_weekly_df.loc[index, 'Regular Time'] = merged_weekly_df.loc[index, 'Lunch Adjusted']
                else:
                    merged_weekly_df.loc[index, 'Overtime'] = merged_weekly_df.loc[index, 'Lunch Adjusted'] - \
                                                              worked_hours_needed
                    merged_weekly_df.loc[index, 'Regular Time'] = worked_hours_needed
                    worked_hours_needed = 0

    # Apply additional checks for errors
    errors_df = merged_weekly_df[(merged_weekly_df['Clock In'].isna()) |
                                 (merged_weekly_df['Clock Out'].isna()) |
                                 ((merged_weekly_df['Lunch Adjusted'] > 8) &
                                  (merged_weekly_df['ApprovedOvertime Start Date'].isnull()) &
                                  (merged_weekly_df['ApprovedOvertime End Date'].isnull())) |
                                 (merged_weekly_df['Lunch Adjusted'] < 8)].copy()
    print("Number of duplicate records:",
          merged_weekly_df.duplicated(subset=['Employee Name', 'Ticket Date', 'JobNo|Customer|Description']).sum())
    duplicates = merged_weekly_df[
        merged_weekly_df.duplicated(subset=['Employee Name', 'Ticket Date', 'JobNo|Customer|Description'], keep=False)]
    print(duplicates)

    # Create the 'Error Description' column
    def generate_error_desc(row):
        if pd.isnull(row['Clock In']) and pd.isnull(row['Clock Out']):
            return 'No Clock In or Clock Out Time'
        elif pd.isnull(row['Clock In']):
            return 'No Clock In'
        elif pd.isnull(row['Clock Out']):
            return 'No Clock Out'
        elif row['Lunch Adjusted'] < 8:
            return 'Less Than 8 Hours'
        else:
            return np.nan


    errors_df['Error Description'] = errors_df.apply(
        generate_error_desc, axis=1)

    # Group the data by 'Employee Name', 'Ticket Date', and 'Day of the Week'
    grouped_df = merged_weekly_df.groupby(
        ['Employee Name', 'Ticket Date', 'Day of the Week'])

    # Calculate the sum of 'Lunch Adjusted' for each group
    merged_weekly_df['Total Lunch Adjusted'] = grouped_df['Lunch Adjusted'].transform(
        'sum')

    # Calculate the cumulative sum of 'Lunch Adjusted' within each group
    merged_weekly_df['Cumulative Lunch Adjusted'] = grouped_df['Lunch Adjusted'].cumsum()

    # Calculate the remaining balance after deducting 40 from 'Cumulative Lunch Adjusted'
    merged_weekly_df['Remaining Balance'] = merged_weekly_df['Cumulative Lunch Adjusted'] - 40

    # Remove the 'Total Lunch Adjusted' column
    merged_weekly_df.drop('Total Lunch Adjusted', axis=1, inplace=True)

    # Convert 'Ticket Date' back to 'mm/dd/yyyy' format
    merged_weekly_df['Ticket Date'] = merged_weekly_df['Ticket Date'].dt.strftime('%m/%d/%Y')
    errors_df['Ticket Date'] = errors_df['Ticket Date'].dt.strftime('%m/%d/%Y')

    # Define the desired column order
    column_order = ['Ticket Date', 'Employee Name', 'Clock In', 'Clock Out', 'Hours Worked',
                    'Lunch Adjusted', 'Regular Time', 'Overtime', 'Day of the Week', 'Employee ID', 'Agency',
                    'Clock-In ID', 'Supervisors Name', 'PM Assigned', 'JobNo|Customer|Description', 'Email',
                    'WTL Approved', 'WTL Start Date', 'WTL End Date', 'ApprovedOvertime', 'ApprovedOvertime Start Date',
                    'ApprovedOvertime End Date']

    # Fill empty Supervisor Name and PM Assigned fields with "NEEDS TO BE ASSIGNED"
    merged_weekly_df['Supervisors Name'] = merged_weekly_df['Supervisors Name'].fillna(
        'NEEDS TO BE ASSIGNED')
    merged_weekly_df['PM Assigned'] = merged_weekly_df['PM Assigned'].fillna(
        'NEEDS TO BE ASSIGNED')

    # Create the 'Error Description' column in errors_df
    errors_df['Error Description'] = errors_df.apply(
        generate_error_desc, axis=1)

    merged_weekly_df['Ticket Date'] = pd.to_datetime(merged_weekly_df['Ticket Date'])
    errors_df['Ticket Date'] = pd.to_datetime(errors_df['Ticket Date'])

    # Convert 'Ticket Date' back to 'mm/dd/yyyy' format in both dataframes
    merged_weekly_df['Ticket Date'] = merged_weekly_df['Ticket Date'].dt.strftime('%m/%d/%Y')
    errors_df['Ticket Date'] = errors_df['Ticket Date'].dt.strftime('%m/%d/%Y')

    # Reorder the columns in the DataFrame
    merged_weekly_df = merged_weekly_df.reindex(columns=column_order)

    # Write the dataframes into a new Excel file with two sheets
    with pd.ExcelWriter('C:/Users/tj-fo/Desktop/Test/PayrollWeekly.xlsx') as writer:
        merged_weekly_df.to_excel(writer, sheet_name='Payroll', index=False)
        errors_df.to_excel(writer, sheet_name='Errors', index=False)

    # Load the workbook
    wb = load_workbook('C:/Users/tj-fo/Desktop/Test/PayrollWeekly.xlsx')

    # Select the sheets
    sheet1 = wb['Payroll']
    sheet2 = wb['Errors']

    # Add new sheets to the workbook
    sheet_names = ["ECO Staffing", "Outsource.net", "Proman Skilled Trades", "Talent Corp", "(blank)"]
    for sheet_name in sheet_names:
        wb.create_sheet(sheet_name)

    # Populate the sheets based on the agency
    for sheet_name in sheet_names:
        sheet = wb[sheet_name]
        if sheet_name == "(blank)":
            # Filter rows where the agency name is null or empty
            blank_agency_rows = merged_weekly_df[
                merged_weekly_df['Agency'].isnull() | merged_weekly_df['Agency'].eq('')]
            blank_agency_data = blank_agency_rows.values.tolist()

            # Write the column names to the sheet
            sheet.append(list(merged_weekly_df.columns))

            # Write the data to the sheet
            for row_data in blank_agency_data:
                sheet.append(row_data)
        else:
            # Filter rows where the agency name matches the sheet name
            agency_rows = merged_weekly_df[merged_weekly_df['Agency'] == sheet_name]
            agency_data = agency_rows.values.tolist()

            # Write the column names to the sheet
            sheet.append(list(merged_weekly_df.columns))

            # Write the data to the sheet
            for row_data in agency_data:
                sheet.append(row_data)

    # Create a red bold font
    red_font = Font(color="FF0000", bold=True)

    # Check each cell in column E (5th column) for both sheets
    for sheet in [sheet1, sheet2]:
        # Modify max_col to 7 for 'Payroll' sheet
        for row in sheet.iter_rows(min_row=2, min_col=4, max_col=7):
            for cell in row:
                if cell.column_letter == 'C' and (cell.value is None or cell.value == ''):
                    cell.value = 'Clock In Time?'
                    cell.font = red_font
                elif cell.column_letter == 'D' and (cell.value is None or cell.value == ''):
                    cell.value = 'Clock Out Time?'
                    cell.font = red_font

    # Apply font color formatting to Overtime column
    for cell in sheet1['H'][1:]:
        overtime_value = cell.value
        approved_start_date = cell.offset(column=13).value
        approved_end_date = cell.offset(column=14).value

        if overtime_value is not None:
            if isinstance(overtime_value, (int, float)):
                if float(overtime_value) > 0 and (approved_start_date is None or approved_end_date is None):
                    cell.font = red_font
                else:
                    cell.font = None
                if float(overtime_value) < 0:
                    cell.value = 0
            else:
                cell.font = red_font
        else:
            cell.font = None

    # Set column widths
    sheet1.column_dimensions['A'].width = 11.26
    sheet1.column_dimensions['B'].width = 26.14
    sheet1.column_dimensions['C'].width = 20
    sheet1.column_dimensions['D'].width = 19
    sheet1.column_dimensions['E'].width = 18
    sheet1.column_dimensions['F'].width = 19
    sheet1.column_dimensions['G'].width = 19
    sheet1.column_dimensions['H'].width = 16
    sheet1.column_dimensions['I'].width = 16
    sheet1.column_dimensions['J'].width = 26
    sheet1.column_dimensions['K'].width = 19
    sheet1.column_dimensions['L'].width = 20
    sheet1.column_dimensions['M'].width = 19
    sheet1.column_dimensions['N'].width = 29
    sheet1.column_dimensions['O'].width = 71.57
    sheet1.column_dimensions['P'].width = 30
    sheet1.column_dimensions['Q'].width = 22
    sheet1.column_dimensions['R'].width = 18
    sheet1.column_dimensions['S'].width = 20.43
    sheet1.column_dimensions['T'].width = 20.43
    sheet1.column_dimensions['U'].width = 29
    sheet1.column_dimensions['V'].width = 29
    sheet1.column_dimensions['W'].width = 29

    # Select the sheets
    sheet2 = wb['Errors']

    # Apply font color formatting to Overtime column (Column G) on the Errors sheet
    for cell in sheet2['G'][1:]:
        overtime_value = cell.value

        if overtime_value is not None:
            if isinstance(overtime_value, (int, float)):
                if float(overtime_value) < 8:
                    cell.font = red_font
            else:
                cell.font = red_font
        else:
            cell.font = None

    # Set column widths
    sheet2.column_dimensions['A'].width = 11.26
    sheet2.column_dimensions['B'].width = 26.14
    sheet2.column_dimensions['C'].width = 31.86
    sheet2.column_dimensions['D'].width = 19
    sheet2.column_dimensions['E'].width = 20.43
    sheet2.column_dimensions['F'].width = 18.71
    sheet2.column_dimensions['G'].width = 20.86
    sheet2.column_dimensions['H'].width = 18
    sheet2.column_dimensions['I'].width = 32.57
    sheet2.column_dimensions['J'].width = 28.71
    sheet2.column_dimensions['K'].width = 20
    sheet2.column_dimensions['L'].width = 22.86
    sheet2.column_dimensions['M'].width = 33.86
    sheet2.column_dimensions['N'].width = 71.57
    sheet2.column_dimensions['O'].width = 31.86
    sheet2.column_dimensions['P'].width = 20.43
    sheet2.column_dimensions['Q'].width = 20.43
    sheet2.column_dimensions['R'].width = 20.43
    sheet2.column_dimensions['S'].width = 20.43
    sheet2.column_dimensions['T'].width = 29
    sheet2.column_dimensions['U'].width = 29
    sheet2.column_dimensions['V'].width = 29
    sheet2.column_dimensions['W'].width = 29

    # Save workbook
    wb.save('C:/Users/tj-fo/Desktop/Test/PayrollWeekly.xlsx')

    # Set all hours as overtime if the day of the week is Saturday or Sunday
    df1.loc[df1['Day of the Week'].isin(
        ['Saturday', 'Sunday']), 'Overtime'] = df1['Lunch Adjusted']

    # Apply additional checks for errors
    errors_df = merged_df[(merged_df['Clock In'].isna()) |
                          (merged_df['Clock Out'].isna()) |
                          ((merged_df['Lunch Adjusted'] > 8) &
                           (merged_df['ApprovedOvertime Start Date'].isnull()) &
                           (merged_df['ApprovedOvertime End Date'].isnull())) |
                          (merged_df['Lunch Adjusted'] < 8)].copy()

    print("Number of duplicate records:",
          merged_df.duplicated(subset=['Employee Name', 'Ticket Date', 'JobNo|Customer|Description']).sum())

    duplicates = merged_df[
        merged_df.duplicated(subset=['Employee Name', 'Ticket Date', 'JobNo|Customer|Description'], keep=False)]
    print(duplicates)

    # Create the 'Error Description' column
    def generate_error_desc(row):
        if pd.isnull(row['Clock In']) and pd.isnull(row['Clock Out']):
            return 'No Clock In or Clock Out Time'
        elif pd.isnull(row['Clock In']):
            return 'No Clock In'
        elif pd.isnull(row['Clock Out']):
            return 'No Clock Out'
        elif row['Lunch Adjusted'] < 8:
            return 'Less Than 8 Hours'
        else:
            return np.nan


    errors_df['Error Description'] = errors_df.apply(
        generate_error_desc, axis=1)

    # Group the data by 'Employee Name', 'Ticket Date', and 'Day of the Week'
    grouped_df = merged_df.groupby(
        ['Employee Name', 'Ticket Date', 'Day of the Week'])

    # Calculate the sum of 'Lunch Adjusted' for each group
    merged_df['Total Lunch Adjusted'] = grouped_df['Lunch Adjusted'].transform(
        'sum')

    # Calculate the cumulative sum of 'Lunch Adjusted' within each group
    merged_df['Cumulative Lunch Adjusted'] = grouped_df['Lunch Adjusted'].cumsum()

    # Calculate the remaining balance after deducting 40 from 'Cumulative Lunch Adjusted'
    merged_df['Remaining Balance'] = merged_df['Cumulative Lunch Adjusted'] - 40

    # Calculate the overtime by subtracting 8 from 'Remaining Balance'
    merged_df['Overtime'] = np.where(
        (merged_df['Remaining Balance'] > 0) & (
            merged_df.duplicated(['Employee Name', 'Ticket Date'])),
        merged_df['Remaining Balance'],
        merged_df['Overtime']
    )

    merged_df['Regular Time'] = np.where(
        (merged_df['Total Lunch Adjusted'] > 8) & (
            merged_df.duplicated(['Employee Name', 'Ticket Date'])),
        0,
        merged_df['Regular Time']
    )

    # Set 'Overtime' equal to 'Lunch Adjusted' for Saturday and Sunday
    merged_df.loc[merged_df['Day of the Week'].isin(
        ['Saturday', 'Sunday']), 'Overtime'] = merged_df['Lunch Adjusted']

    # Set 'Regular Time' to 0 for Saturday and Sunday
    merged_df.loc[merged_df['Day of the Week'].isin(
        ['Saturday', 'Sunday']), 'Regular Time'] = 0

    # Remove the 'Total Lunch Adjusted' column
    merged_df.drop('Total Lunch Adjusted', axis=1, inplace=True)

    # Convert 'Ticket Date' back to 'mm/dd/yyyy' format
    merged_df['Ticket Date'] = merged_df['Ticket Date'].dt.strftime('%m/%d/%Y')
    errors_df['Ticket Date'] = errors_df['Ticket Date'].dt.strftime('%m/%d/%Y')

    # Define the desired column order
    column_order = ['Ticket Date', 'Employee Name', 'Clock In', 'Clock Out', 'Hours Worked',
                    'Lunch Adjusted', 'Regular Time', 'Overtime', 'Day of the Week', 'Employee ID', 'Agency',
                    'Clock-In ID', 'Supervisors Name', 'PM Assigned', 'JobNo|Customer|Description', 'Email',
                    'WTL Approved', 'WTL Start Date', 'WTL End Date', 'ApprovedOvertime', 'ApprovedOvertime Start Date',
                    'ApprovedOvertime End Date']

    # Fill empty Supervisor Name and PM Assigned fields with "NEEDS TO BE ASSIGNED"
    merged_df['Supervisors Name'] = merged_df['Supervisors Name'].fillna(
        'NEEDS TO BE ASSIGNED')
    merged_df['PM Assigned'] = merged_df['PM Assigned'].fillna(
        'NEEDS TO BE ASSIGNED')

    # Create the 'Error Description' column in errors_df
    errors_df['Error Description'] = errors_df.apply(
        generate_error_desc, axis=1)

    merged_df['Ticket Date'] = pd.to_datetime(merged_df['Ticket Date'])
    errors_df['Ticket Date'] = pd.to_datetime(errors_df['Ticket Date'])

    # Convert 'Ticket Date' back to 'mm/dd/yyyy' format in both dataframes
    merged_df['Ticket Date'] = merged_df['Ticket Date'].dt.strftime('%m/%d/%Y')
    errors_df['Ticket Date'] = errors_df['Ticket Date'].dt.strftime('%m/%d/%Y')

    # Reorder the columns in the DataFrame
    merged_df = merged_df.reindex(columns=column_order)

    # Write the dataframes into a new Excel file with two sheets
    with pd.ExcelWriter('C:/Users/tj-fo/Desktop/Test/Payroll.xlsx') as writer:
        merged_df.to_excel(writer, sheet_name='Payroll', index=False)
        errors_df.to_excel(writer, sheet_name='Errors', index=False)

    # Load the workbook
    wb = load_workbook('C:/Users/tj-fo/Desktop/Test/Payroll.xlsx')

    # Select the sheets
    sheet1 = wb['Payroll']
    sheet2 = wb['Errors']

    # Create a red bold font
    red_font = Font(color="FF0000", bold=True)

    # Check each cell in column E (5th column) for both sheets
    for sheet in [sheet1, sheet2]:
        # Modify max_col to 7 for 'Payroll' sheet
        for row in sheet.iter_rows(min_row=2, min_col=4, max_col=7):
            for cell in row:
                if cell.column_letter == 'C' and (cell.value is None or cell.value == ''):
                    cell.value = 'Clock In Time?'
                    cell.font = red_font
                elif cell.column_letter == 'D' and (cell.value is None or cell.value == ''):
                    cell.value = 'Clock Out Time?'
                    cell.font = red_font

    # Apply font color formatting to Overtime column
    for cell in sheet1['H'][1:]:
        overtime_value = cell.value
        approved_start_date = cell.offset(column=13).value
        approved_end_date = cell.offset(column=14).value

        if overtime_value is not None:
            if isinstance(overtime_value, (int, float)):
                if float(overtime_value) > 0 and (approved_start_date is None or approved_end_date is None):
                    cell.font = red_font
                else:
                    cell.font = None
                if float(overtime_value) < 0:
                    cell.value = 0
            else:
                cell.font = red_font
        else:
            cell.font = None

    # Set column widths
    sheet1.column_dimensions['A'].width = 11.26
    sheet1.column_dimensions['B'].width = 26.14
    sheet1.column_dimensions['C'].width = 20
    sheet1.column_dimensions['D'].width = 19
    sheet1.column_dimensions['E'].width = 18
    sheet1.column_dimensions['F'].width = 19
    sheet1.column_dimensions['G'].width = 19
    sheet1.column_dimensions['H'].width = 16
    sheet1.column_dimensions['I'].width = 16
    sheet1.column_dimensions['J'].width = 26
    sheet1.column_dimensions['K'].width = 19
    sheet1.column_dimensions['L'].width = 20
    sheet1.column_dimensions['M'].width = 19
    sheet1.column_dimensions['N'].width = 29
    sheet1.column_dimensions['O'].width = 71.57
    sheet1.column_dimensions['P'].width = 30
    sheet1.column_dimensions['Q'].width = 22
    sheet1.column_dimensions['R'].width = 18
    sheet1.column_dimensions['S'].width = 20.43
    sheet1.column_dimensions['T'].width = 20.43
    sheet1.column_dimensions['U'].width = 29
    sheet1.column_dimensions['V'].width = 29
    sheet1.column_dimensions['W'].width = 29

    # Select the sheets
    sheet2 = wb['Errors']

    # Apply font color formatting to Overtime column (Column G) on the Errors sheet
    for cell in sheet2['G'][1:]:
        overtime_value = cell.value

        if overtime_value is not None:
            if isinstance(overtime_value, (int, float)):
                if float(overtime_value) < 8:
                    cell.font = red_font
            else:
                cell.font = red_font
        else:
            cell.font = None

    # Set column widths
    sheet2.column_dimensions['A'].width = 11.26
    sheet2.column_dimensions['B'].width = 26.14
    sheet2.column_dimensions['C'].width = 31.86
    sheet2.column_dimensions['D'].width = 19
    sheet2.column_dimensions['E'].width = 20.43
    sheet2.column_dimensions['F'].width = 18.71
    sheet2.column_dimensions['G'].width = 20.86
    sheet2.column_dimensions['H'].width = 18
    sheet2.column_dimensions['I'].width = 32.57
    sheet2.column_dimensions['J'].width = 28.71
    sheet2.column_dimensions['K'].width = 20
    sheet2.column_dimensions['L'].width = 22.86
    sheet2.column_dimensions['M'].width = 33.86
    sheet2.column_dimensions['N'].width = 71.57
    sheet2.column_dimensions['O'].width = 31.86
    sheet2.column_dimensions['P'].width = 20.43
    sheet2.column_dimensions['Q'].width = 20.43
    sheet2.column_dimensions['R'].width = 20.43
    sheet2.column_dimensions['S'].width = 20.43
    sheet2.column_dimensions['T'].width = 29
    sheet2.column_dimensions['U'].width = 29
    sheet2.column_dimensions['V'].width = 29
    sheet2.column_dimensions['W'].width = 29

    # Save workbook
    wb.save('C:/Users/tj-fo/Desktop/Test/Payroll.xlsx')

    # Create a new workbook
    insperity_report = Workbook()

    # Create the desired sheets in the workbook
    sheets = ['Hours Import', 'Employee Key', 'Pay Type Key', 'Org Level Items']
    for sheet_name in sheets:
        insperity_report.create_sheet(sheet_name)

    # Delete the default "Sheet"
    default_sheet = insperity_report['Sheet']
    insperity_report.remove(default_sheet)

    # Get the 'Hours Import' sheet from the workbook
    hours_import_sheet = insperity_report['Hours Import']

    # Read the 'PayrollWeekly' workbook
    payroll_weekly = load_workbook('C:/Users/tj-fo/Desktop/Test/PayrollWeekly.xlsx')
    payroll_sheet = payroll_weekly.active

    # Select the desired columns from the 'PayrollWeekly' sheet
    columns_to_copy = ['Employee ID', 'Last Name', 'First Name', 'Ticket Date', 'Regular Time', 'Overtime',
                       'JobNo|Customer|Description']
    # Write the column names in the first row of 'Hours Import' sheet
    for col_idx, column in enumerate(columns_to_copy, start=1):
        hours_import_sheet.cell(row=1, column=col_idx).value = column

    # Copy data from 'PayrollWeekly' to 'Hours Import' sheet
    for row_idx, row in enumerate(payroll_sheet.iter_rows(min_row=2, values_only=True), start=2):
        # Split the 'Employee Name' into 'Last Name' and 'First Name'
        last_name, first_name = row[1].split(',', 1)  # Assuming 'Employee Name' is in the second column (index 1)

        # Write the values to the 'Hours Import' sheet
        hours_import_sheet.cell(row=row_idx, column=1).value = row[9]  # Employee ID (column J)
        hours_import_sheet.cell(row=row_idx, column=2).value = last_name.strip()  # Last Name
        hours_import_sheet.cell(row=row_idx, column=3).value = first_name.strip()  # First Name
        hours_import_sheet.cell(row=row_idx, column=4).value = row[0]  # Ticket Date (column A)
        hours_import_sheet.cell(row=row_idx, column=5).value = row[6]  # Regular Time (column G)
        hours_import_sheet.cell(row=row_idx, column=6).value = row[7]  # Overtime (column H)
        hours_import_sheet.cell(row=row_idx, column=7).value = row[14]  # JobNo|Customer|Description (column O)

    # Apply bold font to the title row in the 'Hours Import' sheet
    title_row = hours_import_sheet[1]
    for cell in title_row:
        cell.font = Font(bold=True)

    # Adjust column widths to fit the contents
    for column in hours_import_sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.coordinate in hours_import_sheet.merged_cells:
                continue
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except TypeError:
                pass
        adjusted_width = (max_length + 2) * 1.2
        hours_import_sheet.column_dimensions[column_letter].width = adjusted_width

    # Save the 'InsperityReport' workbook
    insperity_report.save('C:/Users/tj-fo/Desktop/Test/InsperityReport.xlsx')

except Exception as e:
    print("An error occurred:", str(e))
    raise SystemExit
