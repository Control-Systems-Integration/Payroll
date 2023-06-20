import pandas as pd
import numpy as np
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from openpyxl import load_workbook
from openpyxl.styles import Font

# Create a Tkinter root window
root = Tk()
# Hide the root window
root.withdraw()

try:
    # Open the file picker dialog
    clockIn_File = askopenfilename()
    payRoll_File = askopenfilename()

    # Check if a file path was selected
    if not clockIn_File or not payRoll_File:
        print("No file selected.")
        raise SystemExit

    # Read the Excel files
    df1 = pd.read_excel(clockIn_File)
    df2 = pd.read_excel(payRoll_File)

    # Keep only 'Employee name', 'Ticket Date' and 'Agency' columns in df2
    df2 = df2[
        ['Employee Name', 'Employee ID', 'Ticket Date', 'Agency', 'Clock-In ID', 'Supervisors Name', 'PM Assigned',
         'JobNo|Customer|Description', 'Email', 'WTL Approved', 'WTL Start Date', 'WTL End Date', 'ApprovedOvertime',
         'ApprovedOvertime Start Date', 'ApprovedOvertime End Date']]

    # Convert 'Ticket Date' to datetime in both dataframes
    df1['Ticket Date'] = pd.to_datetime(df1['Ticket Date'])
    df2['Ticket Date'] = pd.to_datetime(df2['Ticket Date'])

    # Calculate 'Lunch Adjusted' as the difference between 'Clock Out' and 'Clock In', converted to hours
    df1['Lunch Adjusted'] = (df1['Clock Out'] - df1['Clock In']).dt.total_seconds() / 3600
    # Taking off the half hour for lunch if Hours Worked is greater than or equal to 5
    df1.loc[df1['Hours Worked'] >= 5, 'Lunch Adjusted'] -= 0.5

    # Add 'Day of the Week' column
    df1['Day of the Week'] = df1['Ticket Date'].dt.day_name()

    # Merge dataframes based on 'Employee name' and 'Ticket Date'
    merged_df = pd.merge(df1, df2, on=['Employee Name', 'Ticket Date', 'JobNo|Customer|Description'], how='left')

    # Remove duplicates from merged_df based on 'Employee name', 'Ticket Date', and 'JobNo|Customer|Description'
    merged_df = merged_df.drop_duplicates(['Employee Name', 'Ticket Date', 'JobNo|Customer|Description'])

    # If 'Agency' is blank, fill with 'CSI'
    merged_df['Agency'] = merged_df['Agency'].fillna('CSI')

    # Calculate Regular Time
    merged_df['Regular Time'] = merged_df['Lunch Adjusted'].where(merged_df['Lunch Adjusted'] <= 8, other=8)

    # Add 0.5 to 'Lunch Adjusted' column if there is a WTL Start Date and WTL End Date
    merged_df.loc[
        ~merged_df['WTL Start Date'].isnull() & ~merged_df['WTL End Date'].isnull(), 'Lunch Adjusted'] += 0.5

    # Calculate Overtime
    merged_df['Overtime'] = merged_df['Lunch Adjusted'] - 8

    # Set all hours as overtime if the day of the week is Saturday or Sunday
    df1.loc[df1['Day of the Week'].isin(['Saturday', 'Sunday']), 'Overtime'] = df1['Lunch Adjusted']

    # Apply additional checks for errors
    errors_df = merged_df[(merged_df['Clock In'].isna()) |
                          (merged_df['Clock Out'].isna()) |
                          ((merged_df['Lunch Adjusted'] > 8) &
                           (merged_df['ApprovedOvertime Start Date'].isnull()) &
                           (merged_df['ApprovedOvertime End Date'].isnull())) |
                          (merged_df['Lunch Adjusted'] < 8)].copy()

    print("Number of duplicate records:",
          merged_df.duplicated(subset=['Employee Name', 'Ticket Date', 'JobNo|Customer|Description']).sum())

    duplicates = merged_df[
        merged_df.duplicated(subset=['Employee Name', 'Ticket Date', 'JobNo|Customer|Description'], keep=False)]
    print(duplicates)

    # Update Overtime for errors
    errors_df.loc[(errors_df['Overtime'] < 0) & (~errors_df['Agency'].str.contains('CSI', case=False)), 'Overtime'] = 0

    # Create the 'Error Description' column
    def generate_error_desc(row):
        if pd.isnull(row['Clock In']) and pd.isnull(row['Clock Out']):
            return 'No Clock In or Clock Out Time'
        elif pd.isnull(row['Clock In']):
            return 'No Clock In'
        elif pd.isnull(row['Clock Out']):
            return 'No Clock Out'
        elif row['Lunch Adjusted'] < 8:
            return 'Less Than 8 Hours'
        else:
            return np.nan

    errors_df['Error Description'] = errors_df.apply(generate_error_desc, axis=1)

    # Group the data by 'Employee Name', 'Ticket Date', and 'Day of the Week'
    grouped_df = merged_df.groupby(['Employee Name', 'Ticket Date', 'Day of the Week'])

    # Calculate the sum of 'Lunch Adjusted' for each group
    merged_df['Total Lunch Adjusted'] = grouped_df['Lunch Adjusted'].transform('sum')

    # Calculate the cumulative sum of 'Lunch Adjusted' within each group
    merged_df['Cumulative Lunch Adjusted'] = grouped_df['Lunch Adjusted'].cumsum()

    # Calculate the remaining balance after deducting 40 from 'Cumulative Lunch Adjusted'
    merged_df['Remaining Balance'] = merged_df['Cumulative Lunch Adjusted'] - 40

    # Calculate the overtime by subtracting 8 from 'Remaining Balance'
    merged_df['Overtime'] = np.where(
        (merged_df['Remaining Balance'] > 0) & (merged_df.duplicated(['Employee Name', 'Ticket Date'])),
        merged_df['Remaining Balance'],
        merged_df['Overtime']
    )

    merged_df['Regular Time'] = np.where(
        (merged_df['Total Lunch Adjusted'] > 8) & (merged_df.duplicated(['Employee Name', 'Ticket Date'])),
        0,
        merged_df['Regular Time']
    )

    # Set 'Overtime' equal to 'Lunch Adjusted' for Saturday and Sunday
    merged_df.loc[merged_df['Day of the Week'].isin(['Saturday', 'Sunday']), 'Overtime'] = merged_df['Lunch Adjusted']

    # Set 'Regular Time' to 0 for Saturday and Sunday
    merged_df.loc[merged_df['Day of the Week'].isin(['Saturday', 'Sunday']), 'Regular Time'] = 0

    # Remove the 'Total Lunch Adjusted' column
    merged_df.drop('Total Lunch Adjusted', axis=1, inplace=True)

    # Convert 'Ticket Date' back to 'mm/dd/yyyy' format
    merged_df['Ticket Date'] = merged_df['Ticket Date'].dt.strftime('%m/%d/%Y')
    errors_df['Ticket Date'] = errors_df['Ticket Date'].dt.strftime('%m/%d/%Y')

    # Define the desired column order
    column_order = ['Ticket Date', 'Employee Name', 'Clock In', 'Clock Out', 'Hours Worked',
                    'Lunch Adjusted', 'Regular Time', 'Overtime', 'Day of the Week', 'Employee ID', 'Agency',
                    'Clock-In ID','Supervisors Name', 'PM Assigned', 'JobNo|Customer|Description', 'Email',
                    'WTL Approved', 'WTL Start Date', 'WTL End Date', 'ApprovedOvertime', 'ApprovedOvertime Start Date',
                    'ApprovedOvertime End Date']

    # Fill empty Supervisor Name and PM Assigned fields with "NEEDS TO BE ASSIGNED"
    merged_df['Supervisors Name'] = merged_df['Supervisors Name'].fillna('NEEDS TO BE ASSIGNED')
    merged_df['PM Assigned'] = merged_df['PM Assigned'].fillna('NEEDS TO BE ASSIGNED')

    # Create the 'Error Description' column in errors_df
    errors_df['Error Description'] = errors_df.apply(generate_error_desc, axis=1)

    merged_df['Ticket Date'] = pd.to_datetime(merged_df['Ticket Date'])
    errors_df['Ticket Date'] = pd.to_datetime(errors_df['Ticket Date'])

    # Convert 'Ticket Date' back to 'mm/dd/yyyy' format in both dataframes
    merged_df['Ticket Date'] = merged_df['Ticket Date'].dt.strftime('%m/%d/%Y')
    errors_df['Ticket Date'] = errors_df['Ticket Date'].dt.strftime('%m/%d/%Y')

    # Reorder the columns in the DataFrame
    merged_df = merged_df.reindex(columns=column_order)

    # Write the dataframes into a new Excel file with two sheets
    with pd.ExcelWriter('C:/Users/tj-fo/Desktop/Test/Payroll.xlsx') as writer:
        merged_df.to_excel(writer, sheet_name='Payroll', index=False)
        errors_df.to_excel(writer, sheet_name='Errors', index=False)

    # Load the workbook
    wb = load_workbook('C:/Users/tj-fo/Desktop/Test/Payroll.xlsx')

    # Select the sheets
    sheet1 = wb['Payroll']
    sheet2 = wb['Errors']

    # Create a red bold font
    red_font = Font(color="FF0000", bold=True)

    # Check each cell in column E (5th column) for both sheets
    for sheet in [sheet1, sheet2]:
        for row in sheet.iter_rows(min_row=2, min_col=4, max_col=7):  # Modify max_col to 7 for 'Payroll' sheet
            for cell in row:
                if cell.column_letter == 'C' and (cell.value is None or cell.value == ''):
                    cell.value = 'Clock In Time?'
                    cell.font = red_font
                elif cell.column_letter == 'D' and (cell.value is None or cell.value == ''):
                    cell.value = 'Clock Out Time?'
                    cell.font = red_font

    # Apply font color formatting to Overtime column
    for cell in sheet1['H'][1:]:
        overtime_value = cell.value
        approved_start_date = cell.offset(column=13).value
        approved_end_date = cell.offset(column=14).value

        if overtime_value is not None:
            if isinstance(overtime_value, (int, float)):
                if float(overtime_value) > 0 and (approved_start_date is None or approved_end_date is None):
                    cell.font = red_font
                else:
                    cell.font = None
                if float(overtime_value) < 0:
                    cell.value = 0
            else:
                cell.font = red_font
        else:
            cell.font = None

    # Set column widths
    sheet1.column_dimensions['A'].width = 11.26
    sheet1.column_dimensions['B'].width = 26.14
    sheet1.column_dimensions['C'].width = 20
    sheet1.column_dimensions['D'].width = 19
    sheet1.column_dimensions['E'].width = 18
    sheet1.column_dimensions['F'].width = 19
    sheet1.column_dimensions['G'].width = 19
    sheet1.column_dimensions['H'].width = 16
    sheet1.column_dimensions['I'].width = 16
    sheet1.column_dimensions['J'].width = 26
    sheet1.column_dimensions['K'].width = 19
    sheet1.column_dimensions['L'].width = 20
    sheet1.column_dimensions['M'].width = 19
    sheet1.column_dimensions['N'].width = 29
    sheet1.column_dimensions['O'].width = 71.57
    sheet1.column_dimensions['P'].width = 30
    sheet1.column_dimensions['Q'].width = 22
    sheet1.column_dimensions['R'].width = 18
    sheet1.column_dimensions['S'].width = 20.43
    sheet1.column_dimensions['T'].width = 20.43
    sheet1.column_dimensions['U'].width = 29
    sheet1.column_dimensions['V'].width = 29
    sheet1.column_dimensions['W'].width = 29

    # Select the sheets
    sheet2 = wb['Errors']

    # Apply font color formatting to Overtime column (Column G) on the Errors sheet
    for cell in sheet2['G'][1:]:
        overtime_value = cell.value

        if overtime_value is not None:
            if isinstance(overtime_value, (int, float)):
                if float(overtime_value) < 8:
                    cell.font = red_font
            else:
                cell.font = red_font
        else:
            cell.font = None

    # Set column widths
    sheet2.column_dimensions['A'].width = 11.26
    sheet2.column_dimensions['B'].width = 26.14
    sheet2.column_dimensions['C'].width = 31.86
    sheet2.column_dimensions['D'].width = 19
    sheet2.column_dimensions['E'].width = 20.43
    sheet2.column_dimensions['F'].width = 18.71
    sheet2.column_dimensions['G'].width = 20.86
    sheet2.column_dimensions['H'].width = 18
    sheet2.column_dimensions['I'].width = 32.57
    sheet2.column_dimensions['J'].width = 28.71
    sheet2.column_dimensions['K'].width = 20
    sheet2.column_dimensions['L'].width = 22.86
    sheet2.column_dimensions['M'].width = 33.86
    sheet2.column_dimensions['N'].width = 71.57
    sheet2.column_dimensions['O'].width = 31.86
    sheet2.column_dimensions['P'].width = 20.43
    sheet2.column_dimensions['Q'].width = 20.43
    sheet2.column_dimensions['R'].width = 20.43
    sheet2.column_dimensions['S'].width = 20.43
    sheet2.column_dimensions['T'].width = 29
    sheet2.column_dimensions['U'].width = 29
    sheet2.column_dimensions['V'].width = 29
    sheet2.column_dimensions['W'].width = 29

    # Save workbook
    wb.save('C:/Users/tj-fo/Desktop/Test/Payroll.xlsx')

    # Load the Excel file
    df = pd.read_excel('C:/Users/tj-fo/Desktop/Test/Payroll.xlsx')

    # Convert the 'Ticket Date', 'Clock In', and 'Clock Out' columns to datetime
    df['Ticket Date'] = pd.to_datetime(df['Ticket Date'])
    df['Clock In'] = pd.to_datetime(df['Clock In'])
    df['Clock Out'] = pd.to_datetime(df['Clock Out'])

    # Remove the calculation of total hours worked
    df['Total Hours Worked'] = df['Lunch Adjusted']

    # Create a list to hold the results
    results = []

    # Create a DataFrame to hold missing Employees
    missing_employees = pd.DataFrame(columns=['Employee Name', 'Ticket Date'])

    # Group by 'Employee Name', 'JobNo|Customer|Description', 'Agency', and 'Ticket Date'
    grouped = df.groupby(['Employee Name', 'JobNo|Customer|Description', 'Agency', df['Ticket Date'].dt.date])

    for name, group in grouped:
        total_hours = group['Total Hours Worked'].sum()
        overtime_hours = group['Overtime'].sum()

        results.append(pd.DataFrame({
            'Employee Name': [name[0]],
            'Employee ID': [group['Employee ID'].iloc[0]],
            'JobNo|Customer|Description': [name[1]],
            'Agency': [name[2]],
            'Ticket Date': [name[3]],
            'Day': [group['Ticket Date'].dt.day_name().iloc[0]],
            'Regular Hours': [total_hours],
            'Overtime Hours': [overtime_hours],
            'Clock In': [group['Clock In'].iloc[0]],
            'Clock Out': [group['Clock Out'].iloc[0]],
            'Supervisors Name': [group['Supervisors Name'].iloc[0]],
            'PM Assigned': [group['PM Assigned'].iloc[0]],
            'Email': [group['Email'].iloc[0]],
            'WTL Approved': [group['WTL Approved'].iloc[0]],
            'ApprovedOvertime': [group['ApprovedOvertime'].iloc[0]]
        }))

    # Concatenate all the results into a single dataframe
    result = pd.concat(results)

    # Save the result to a new Excel file
    result.to_excel('C:/Users/tj-fo/Desktop/Test/Results.xlsx', index=False)

    # Load the workbook
    book = load_workbook('C:/Users/tj-fo/Desktop/Test/Results.xlsx')

    # Access the sheet by name or index
    sheet1 = book['Sheet1']

    # Set column widths
    sheet1.column_dimensions['A'].width = 29.71
    sheet1.column_dimensions['B'].width = 12.57
    sheet1.column_dimensions['C'].width = 72.71
    sheet1.column_dimensions['D'].width = 17
    sheet1.column_dimensions['E'].width = 14.29
    sheet1.column_dimensions['F'].width = 15.14
    sheet1.column_dimensions['G'].width = 15
    sheet1.column_dimensions['H'].width = 16.71
    sheet1.column_dimensions['I'].width = 20.29
    sheet1.column_dimensions['J'].width = 27.71
    sheet1.column_dimensions['K'].width = 22.43
    sheet1.column_dimensions['L'].width = 23.57
    sheet1.column_dimensions['M'].width = 35.29
    sheet1.column_dimensions['N'].width = 20.57
    sheet1.column_dimensions['O'].width = 20.57

    # Save the modified workbook
    book.save('C:/Users/tj-fo/Desktop/Test/Results.xlsx')
except Exception as e:
    print("An error occurred:", str(e))
    raise SystemExit
