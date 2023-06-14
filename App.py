import datetime
import pandas as pd
import numpy as np
from tkinter import Tk
from tkinter.filedialog import askopenfilename, asksaveasfilename
from openpyxl import load_workbook
from openpyxl.styles import Font, Color

# Create a Tkinter root window
root = Tk()
# Hide the root window
root.withdraw()

try:
    # Open the file picker dialog
    clockIn_File = askopenfilename()
    payRoll_File = askopenfilename()

    # Check if a file path was selected
    if not clockIn_File or not payRoll_File:
        print("No file selected.")
        raise SystemExit

    # Read the Excel files
    df1 = pd.read_excel(clockIn_File)
    df2 = pd.read_excel(payRoll_File)

    # Keep only 'Employee name', 'Ticket Date' and 'Agency' columns in df2
    df2 = df2[
        ['Employee Name', 'Employee ID', 'Ticket Date', 'Agency', 'Clock-In ID', 'Supervisors Name', 'PM Assigned',
         'JobNo|Customer|Description', 'Email', 'WTL Approved', 'WTL Start Date', 'WTL End Date', 'ApprovedOvertime',
         'ApprovedOvertime Start Date', 'ApprovedOvertime End Date']]

    # Convert 'Ticket Date' to datetime in both dataframes
    df1['Ticket Date'] = pd.to_datetime(df1['Ticket Date'])
    df2['Ticket Date'] = pd.to_datetime(df2['Ticket Date'])

    # If 'Clock In' and 'Clock Out' are not datetime, convert them
    df1['Clock In'] = pd.to_datetime(df1['Clock In'])
    df1['Clock Out'] = pd.to_datetime(df1['Clock Out'])

    # Calculate 'Actual Hours Worked' as the difference between 'Clock Out' and 'Clock In', converted to hours
    df1['Actual Hours Worked'] = (df1['Clock Out'] - df1['Clock In']).dt.total_seconds() / 3600
    # Taking off the half hour for lunch
    df1['Actual Hours Worked'] = (df1['Actual Hours Worked'] - .5)

    # Add 'Day of the Week' column
    df1['Day of the Week'] = df1['Ticket Date'].dt.day_name()

    # Merge dataframes based on 'Employee name' and 'Ticket Date'
    merged_df = pd.merge(df1, df2, on=['Employee Name', 'Ticket Date'], how='left')

    # Add 0.5 to 'Actual Hours Worked' column if there is a WTL Start Date and WTL End Date
    merged_df.loc[
        ~merged_df['WTL Start Date'].isnull() & ~merged_df['WTL End Date'].isnull(), 'Actual Hours Worked'] += 0.5

    approved_overtime_mask = (~merged_df['ApprovedOvertime Start Date'].isnull()) & \
                             (~merged_df['ApprovedOvertime End Date'].isnull()) & \
                             (merged_df['Actual Hours Worked'] > 8)

    merged_df.loc[approved_overtime_mask, 'Overtime'] = merged_df.loc[approved_overtime_mask, 'Actual Hours Worked'] - 8
    merged_df.loc[~approved_overtime_mask, 'Overtime'] = 0

    # If 'Agency' is blank, fill with 'CSI'
    merged_df['Agency'] = merged_df['Agency'].fillna('CSI')

    errors_df = merged_df[(merged_df['Clock In'].isna()) |
                          (merged_df['Clock Out'].isna()) |
                          ((merged_df['Actual Hours Worked'] > 8) &
                           (merged_df['ApprovedOvertime Start Date'].isnull()) &
                           (merged_df['ApprovedOvertime End Date'].isnull())) |
                          (merged_df['Actual Hours Worked'] < 8)].copy()

    # Create the 'Error Description' column
    def generate_error_desc(row):
        if pd.isnull(row['Clock In']) and pd.isnull(row['Clock Out']):
            return 'No Clock In or Clock Out Time'
        elif pd.isnull(row['Clock In']):
            return 'No Clock In'
        elif pd.isnull(row['Clock Out']):
            return 'No Clock Out'
        elif row['Actual Hours Worked'] < 8:
            return 'Less Than 8 Hours'
        else:
            return np.nan


    errors_df['Error Description'] = errors_df.apply(generate_error_desc, axis=1)

    # Convert 'Ticket Date' back to 'mm/dd/yyyy' format
    merged_df['Ticket Date'] = merged_df['Ticket Date'].dt.strftime('%m/%d/%Y')
    errors_df['Ticket Date'] = errors_df['Ticket Date'].dt.strftime('%m/%d/%Y')

    # Define the desired column order
    column_order = ['Ticket Date', 'Quote/Job Number Number', 'Employee Name', 'Clock In', 'Clock Out', 'Hours Worked',
                    'Actual Hours Worked', 'Overtime', 'Day of the Week', 'Employee ID', 'Agency', 'Clock-In ID',
                    'Supervisors Name', 'PM Assigned', 'JobNo|Customer|Description', 'Email', 'WTL Approved',
                    'WTL Start Date', 'WTL End Date', 'ApprovedOvertime', 'ApprovedOvertime Start Date',
                    'ApprovedOvertime End Date']

    # Reorder the columns in the DataFrame
    merged_df = merged_df.reindex(columns=column_order)

    # Write the dataframes into a new Excel file with two sheets
    with pd.ExcelWriter('C:/Users/tj-fo/Desktop/Test/Payroll.xlsx') as writer:
        merged_df.to_excel(writer, sheet_name='Payroll', index=False)
        errors_df.to_excel(writer, sheet_name='Errors', index=False)

    # SB
    # Load the workbook
    wb = load_workbook('C:/Users/tj-fo/Desktop/Test/Payroll.xlsx')

    # SB
    # Select the sheets
    sheet1 = wb['Payroll']

    # SB
    # Create a red bold font
    red_bold_font = Font(color="FF0000", bold=True)

    # Check each cell in column E (5th column) for both sheets
    for sheet in [sheet1]:
        for row in sheet.iter_rows(min_row=2, min_col=4, max_col=5):
            for cell in row:
                if cell.column_letter == 'D' and (cell.value is None or cell.value == ''):
                    cell.value = 'Clock In Time?'
                    cell.font = red_bold_font
                elif cell.column_letter == 'E' and (cell.value is None or cell.value == ''):
                    cell.value = 'Clock Out Time?'
                    cell.font = red_bold_font

    # SB
    # Set column widths
    sheet1.column_dimensions['A'].width = 11.26
    sheet1.column_dimensions['B'].width = 26.14
    sheet1.column_dimensions['C'].width = 31.86
    sheet1.column_dimensions['D'].width = 19
    sheet1.column_dimensions['E'].width = 20.43
    sheet1.column_dimensions['F'].width = 18.71
    sheet1.column_dimensions['G'].width = 20.86
    sheet1.column_dimensions['H'].width = 16
    sheet1.column_dimensions['I'].width = 26
    sheet1.column_dimensions['J'].width = 19
    sheet1.column_dimensions['K'].width = 20
    sheet1.column_dimensions['L'].width = 19
    sheet1.column_dimensions['M'].width = 29
    sheet1.column_dimensions['N'].width = 22
    sheet1.column_dimensions['O'].width = 71.57
    sheet1.column_dimensions['P'].width = 25
    sheet1.column_dimensions['Q'].width = 18
    sheet1.column_dimensions['R'].width = 20.43
    sheet1.column_dimensions['S'].width = 20.43
    sheet1.column_dimensions['T'].width = 29
    sheet1.column_dimensions['U'].width = 29
    sheet1.column_dimensions['V'].width = 29

    # Select the sheets
    sheet2 = wb['Errors']

    # Set column widths
    sheet2.column_dimensions['A'].width = 11.26
    sheet2.column_dimensions['B'].width = 26.14
    sheet2.column_dimensions['C'].width = 31.86
    sheet2.column_dimensions['D'].width = 19
    sheet2.column_dimensions['E'].width = 20.43
    sheet2.column_dimensions['F'].width = 18.71
    sheet2.column_dimensions['G'].width = 20.86
    sheet2.column_dimensions['H'].width = 18
    sheet2.column_dimensions['I'].width = 32.57
    sheet2.column_dimensions['J'].width = 28.71
    sheet2.column_dimensions['K'].width = 20
    sheet2.column_dimensions['L'].width = 22.86
    sheet2.column_dimensions['M'].width = 33.86
    sheet2.column_dimensions['N'].width = 71.57
    sheet2.column_dimensions['O'].width = 31.86
    sheet2.column_dimensions['P'].width = 20.43
    sheet2.column_dimensions['Q'].width = 20.43
    sheet2.column_dimensions['R'].width = 20.43

    # Save workbook
    wb.save('C:/Users/tj-fo/Desktop/Test/Payroll.xlsx')


    def round_time(dt):
        # Calculate the number of minutes past the last 15-minute mark
        minutes = (dt.minute % 15) * 60 + dt.second

        # If the number of minutes is less than 7, round down; otherwise, round up
        if minutes < 7 * 60:
            dt = dt - datetime.timedelta(minutes=dt.minute % 15, seconds=dt.second)
        else:
            dt = dt + datetime.timedelta(minutes=15 - dt.minute % 15, seconds=-dt.second)

        return dt


    # Load the Excel file
    df = pd.read_excel('C:/Users/tj-fo/Desktop/Test/Payroll.xlsx')

    # Load the Employee Name data from Test3.xlsx file
    df_test3 = pd.read_excel('C:/Users/tj-fo/Desktop/Test/Test3.xlsx')
    employee_names_test3 = df_test3['Employee Name'].unique()

    # Convert the 'Ticket Date', 'Clock In', and 'Clock Out' columns to datetime
    df['Ticket Date'] = pd.to_datetime(df['Ticket Date'])
    df['Clock In'] = pd.to_datetime(df['Clock In']).apply(round_time)
    df['Clock Out'] = pd.to_datetime(df['Clock Out']).apply(round_time)

    # Calculate the total hours worked for each job
    df['Total Hours Worked'] = (df['Clock Out'] - df['Clock In']).dt.total_seconds() / 3600

    # Create a list to hold the results
    results = []

    # Create a DataFrame to hold missing Employees
    missing_employees = pd.DataFrame(columns=['Employee Name', 'Ticket Date'])

    # Group by 'Employee Name', 'JobNo|Customer|Description', 'Agency', and 'Ticket Date'
    grouped = df.groupby(['Employee Name', 'JobNo|Customer|Description', 'Agency', df['Ticket Date'].dt.date])

    for name, group in grouped:
        if name[0] not in employee_names_test3:
            missing_employees = missing_employees.append({'Employee Name': name[0], 'Ticket Date': name[3]},
                                                         ignore_index=True)

    for name, group in grouped:
        total_hours = group['Total Hours Worked'].sum()

        # Deduct 30 minutes for lunch break if the employee worked for more than 5 hours
        if total_hours > 5:
            total_hours -= 0.5

        regular_hours = round(min(8, total_hours), 2)
        overtime_hours = round(max(0, total_hours - 8), 2)

        # SB
        print(group.columns)

        # If the Ticket Date is on a Saturday or Sunday, all hours are overtime
        if group['Ticket Date'].dt.dayofweek.iloc[0] >= 5:
            overtime_hours = round(total_hours, 2)
            regular_hours = 0

        results.append(pd.DataFrame({
            'Employee Name': [name[0]],
            'Employee ID': [group['Employee ID'].iloc[0]],
            'JobNo|Customer|Description': [name[1]],
            'Agency': [name[2]],
            'Ticket Date': [name[3]],
            'Day': [group['Ticket Date'].dt.day_name().iloc[0]],
            'Regular Hours': [regular_hours],
            'Overtime Hours': [overtime_hours],
            'Clock In': [group['Clock In'].iloc[0]],
            'Clock Out': [group['Clock Out'].iloc[0]],
            'Supervisors Name': [group['Supervisors Name'].iloc[0]],
            'PM Assigned': [group['PM Assigned'].iloc[0]],
            'Email': [group['Email'].iloc[0]],
            'WTL Approved': [group['WTL Approved'].iloc[0]],
            'ApprovedOvertime': [group['ApprovedOvertime'].iloc[0]]
        }))

    # Concatenate all the results into a single dataframe
    result = pd.concat(results)

    # Save the result to a new Excel file
    result.to_excel('C:/Users/tj-fo/Desktop/Test/Results.xlsx', index=False)

    # Load the workbook
    book = load_workbook('C:/Users/tj-fo/Desktop/Test/Results.xlsx')

    # Access the sheet by name or index
    sheet1 = book['Sheet1']

    # Set column widths
    sheet1.column_dimensions['A'].width = 29.71
    sheet1.column_dimensions['B'].width = 12.57
    sheet1.column_dimensions['C'].width = 72.71
    sheet1.column_dimensions['D'].width = 17
    sheet1.column_dimensions['E'].width = 14.29
    sheet1.column_dimensions['F'].width = 15.14
    sheet1.column_dimensions['G'].width = 15
    sheet1.column_dimensions['H'].width = 16.71
    sheet1.column_dimensions['I'].width = 20.29
    sheet1.column_dimensions['J'].width = 27.71
    sheet1.column_dimensions['K'].width = 22.43
    sheet1.column_dimensions['L'].width = 23.57
    sheet1.column_dimensions['M'].width = 35.29
    sheet1.column_dimensions['N'].width = 20.57

    # Save the modified workbook
    book.save('C:/Users/tj-fo/Desktop/Test/Results.xlsx')
except Exception as e:
    print("An error occurred:", str(e))
    raise SystemExit
